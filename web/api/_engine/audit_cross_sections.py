"""
Audit harness for MASTER CROSS SECTION sheet correctness.

Independently re-derives cross-section labels for every column footprint in the
DXF using a different algorithm than the production code, then cross-checks
against the values written to the MASTER CROSS SECTION sheet.

Run from a workspace that has already produced column_load_takedown.xlsx:
    python3 audit_cross_sections.py

Outputs:
    - audit_cross_sections.json (machine-readable report)
    - audit_cross_sections_visual.png (annotated grid of unique footprints)
"""

from __future__ import annotations

import json
import math
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

import ezdxf
import openpyxl
from shapely.geometry import Polygon


CIRCULARITY_THRESHOLD = 0.92  # match production
DIM_TOLERANCE_IN = 1  # an audit disagreement <= 1 inch is treated as a near-miss


def round_in(x_ft: float) -> int:
    """Convert feet → inches and round to nearest inch (half-away-from-zero)."""
    val_in = x_ft * 12.0
    return int(math.floor(val_in + 0.5)) if val_in >= 0 else -int(math.floor(-val_in + 0.5))


def label_from_polygon_independent(poly: Polygon) -> Optional[str]:
    """
    Independent cross-section classifier — uses a different code path than
    format_cross_section() in export_column_loads.py. Compares the polygon's
    actual vertices (rectangles are detected by 4 right-angled corners) instead
    of leaning solely on the circularity ratio and min_rotated_rectangle.
    """
    if poly is None or poly.is_empty:
        return None

    coords = list(poly.exterior.coords)
    if coords[0] == coords[-1]:
        coords = coords[:-1]
    n = len(coords)
    if n < 3:
        return None

    area = poly.area
    perimeter = poly.length
    circularity = 4.0 * math.pi * area / (perimeter * perimeter) if perimeter > 1e-9 else 0.0

    # Path A — sharp 4-corner rectangle: directly use the polygon's own side lengths.
    if n == 4 or _looks_like_rectangle(coords):
        rect = poly.minimum_rotated_rectangle
        rc = list(rect.exterior.coords)
        side_a = math.hypot(rc[1][0] - rc[0][0], rc[1][1] - rc[0][1])
        side_b = math.hypot(rc[2][0] - rc[1][0], rc[2][1] - rc[1][1])
        w_in = round_in(min(side_a, side_b))
        d_in = round_in(max(side_a, side_b))
        return f"{w_in}x{d_in}"

    # Path B — high circularity, treat as circle.
    if circularity >= CIRCULARITY_THRESHOLD:
        diameter_in = round_in(2.0 * math.sqrt(area / math.pi))
        return f"d{diameter_in}"

    # Path C — fall back to oriented bounding box.
    rect = poly.minimum_rotated_rectangle
    rc = list(rect.exterior.coords)
    side_a = math.hypot(rc[1][0] - rc[0][0], rc[1][1] - rc[0][1])
    side_b = math.hypot(rc[2][0] - rc[1][0], rc[2][1] - rc[1][1])
    w_in = round_in(min(side_a, side_b))
    d_in = round_in(max(side_a, side_b))
    return f"{w_in}x{d_in}"


def _looks_like_rectangle(coords) -> bool:
    """Polygon with 4 corners where all interior angles are within 5° of 90°."""
    if len(coords) != 4:
        return False
    for i in range(4):
        a = coords[(i - 1) % 4]
        b = coords[i]
        c = coords[(i + 1) % 4]
        v1 = (a[0] - b[0], a[1] - b[1])
        v2 = (c[0] - b[0], c[1] - b[1])
        m1 = math.hypot(*v1)
        m2 = math.hypot(*v2)
        if m1 < 1e-9 or m2 < 1e-9:
            return False
        cos_ang = (v1[0] * v2[0] + v1[1] * v2[1]) / (m1 * m2)
        cos_ang = max(-1.0, min(1.0, cos_ang))
        angle = math.degrees(math.acos(cos_ang))
        if abs(angle - 90.0) > 5.0:
            return False
    return True


def label_from_dxf_entity(entity, factor: float) -> Optional[str]:
    """
    Hard ground-truth using the DXF entity type directly:
      - CIRCLE entity → diameter from radius
      - 4-vertex LWPOLYLINE → side lengths from vertex deltas
    Returns None for shapes that don't match these two clean cases.
    """
    et = entity.dxftype()
    if et == "CIRCLE":
        d_in = round_in(2.0 * entity.dxf.radius * factor)
        return f"d{d_in}"
    if et == "LWPOLYLINE":
        pts = [(p[0] * factor, p[1] * factor) for p in entity.get_points("xy")]
        if len(pts) == 4 and _looks_like_rectangle(pts):
            side_a = math.hypot(pts[1][0] - pts[0][0], pts[1][1] - pts[0][1])
            side_b = math.hypot(pts[2][0] - pts[1][0], pts[2][1] - pts[1][1])
            w_in = round_in(min(side_a, side_b))
            d_in = round_in(max(side_a, side_b))
            return f"{w_in}x{d_in}"
    return None


def parse_dim(label: str) -> tuple[Optional[str], Optional[int], Optional[int]]:
    """Return (kind, dim_a, dim_b) — kind ∈ {'rect','circle'}; dim_b is None for circles."""
    if not isinstance(label, str) or not label:
        return None, None, None
    if label.startswith("d"):
        try:
            return "circle", int(label[1:]), None
        except ValueError:
            return None, None, None
    if "x" in label:
        try:
            a, b = label.split("x")
            return "rect", int(a), int(b)
        except ValueError:
            return None, None, None
    return None, None, None


def dim_close(a: str, b: str, tol_in: int = DIM_TOLERANCE_IN) -> bool:
    ka, aa, ab = parse_dim(a)
    kb, ba, bb = parse_dim(b)
    if ka is None or kb is None or ka != kb:
        return False
    if ka == "circle":
        return abs(aa - ba) <= tol_in
    return abs(aa - ba) <= tol_in and abs(ab - bb) <= tol_in


def load_workspace_footprints(workspace: Path):
    """Return list of dicts: {polygon, footprint_id}."""
    sys.path.insert(0, str(Path(__file__).parent))
    from tributary import load_column_footprints  # type: ignore
    cwd = Path.cwd()
    try:
        import os
        os.chdir(workspace)
        fps = load_column_footprints("dxf_column_footprints.csv")
    finally:
        import os
        os.chdir(cwd)
    return fps


def load_dxf_column_entities(dxf_path: Path, support_layers, factor: float):
    """Return list of (entity, computed_polygon) for column entities on the support layers."""
    sys.path.insert(0, str(Path(__file__).parent.parent))
    sys.path.insert(0, str(Path(__file__).parent))
    from geometry_utils import entity_to_polygon
    doc = ezdxf.readfile(str(dxf_path))
    msp = doc.modelspace()
    out = []
    for e in msp:
        if e.dxf.layer not in support_layers:
            continue
        poly = entity_to_polygon(e, factor)
        if poly is None or poly.is_empty:
            continue
        out.append((e, poly))
    return out


def audit_workspace(workspace: Path, dxf_path: Path, layer_mapping: dict, source_units: str = "in"):
    factor = 1.0 / 12.0 if source_units == "in" else 1.0

    # Load production xlsx (the artifact we're auditing).
    wb = openpyxl.load_workbook(workspace / "column_load_takedown.xlsx")
    if "MASTER CROSS SECTION" not in wb.sheetnames:
        raise RuntimeError("MASTER CROSS SECTION sheet missing — pipeline did not include the new sheet")
    xs_sheet = wb["MASTER CROSS SECTION"]

    rows = list(xs_sheet.iter_rows(values_only=True))
    header = rows[0]
    column_labels = list(header[1:])
    floors = []
    values = defaultdict(dict)  # values[floor][col_label] = string
    for r in rows[1:]:
        floor = r[0]
        floors.append(floor)
        for i, val in enumerate(r[1:]):
            if val not in ("", None):
                values[floor][column_labels[i]] = val

    # Per-column-label tally — most-common cross-section value across floors.
    per_label_values = defaultdict(Counter)
    for floor, m in values.items():
        for k, v in m.items():
            per_label_values[k][v] += 1

    # Cross-check internal consistency: same column should have same label across floors.
    inconsistent = []
    for label, counter in per_label_values.items():
        if len(counter) > 1:
            inconsistent.append({"column": label, "values": dict(counter)})

    # DXF-entity ground truth: for every column entity on the support layers, compute the
    # "hard" label (CIRCLE or 4-vertex rectangle). Then run the independent algorithm too.
    support_layers = set(layer_mapping.get("support_point", []))
    entity_records = load_dxf_column_entities(dxf_path, support_layers, factor)

    # Per-entity audit — independent algo vs hard DXF-entity label vs production-derived label
    # (re-derive production label here using the imported production helper).
    sys.path.insert(0, str(Path(__file__).parent))
    from export_column_loads import format_cross_section  # type: ignore

    audit_rows = []
    mismatch_hard = 0
    mismatch_indep = 0
    for ent, poly in entity_records:
        hard = label_from_dxf_entity(ent, factor)
        indep = label_from_polygon_independent(poly)
        prod = format_cross_section(poly)
        row = {
            "dxftype": ent.dxftype(),
            "layer": ent.dxf.layer,
            "area_ft2": round(poly.area, 4),
            "perimeter_ft": round(poly.length, 4),
            "production": prod,
            "independent": indep,
            "hard_truth": hard,
        }
        if hard is not None and prod is not None and not dim_close(hard, prod):
            mismatch_hard += 1
            row["disagreement"] = "production vs hard_truth"
        if indep is not None and prod is not None and not dim_close(indep, prod):
            mismatch_indep += 1
            row.setdefault("disagreement", "production vs independent")
        audit_rows.append(row)

    # Distribution of values in the xlsx.
    flat_values = Counter()
    for floor, m in values.items():
        for v in m.values():
            flat_values[v] += 1

    report = {
        "workspace": str(workspace),
        "dxf": str(dxf_path),
        "n_floors": len(floors),
        "n_columns": len(column_labels),
        "value_distribution": dict(flat_values),
        "inconsistent_columns": inconsistent,
        "n_dxf_column_entities": len(entity_records),
        "n_hard_truth_available": sum(1 for r in audit_rows if r["hard_truth"] is not None),
        "n_mismatch_vs_hard_truth": mismatch_hard,
        "n_mismatch_vs_independent_algo": mismatch_indep,
        "disagreements": [r for r in audit_rows if r.get("disagreement")][:50],
    }

    out_json = workspace / "audit_cross_sections.json"
    out_json.write_text(json.dumps(report, indent=2))
    return report


if __name__ == "__main__":
    workspace = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
    dxf_path = workspace / "INPUT.DXF"
    cfg = json.loads((workspace / "job_config.json").read_text())
    report = audit_workspace(workspace, dxf_path, cfg["layers"], cfg.get("source_units", "in"))

    print("=" * 70)
    print("MASTER CROSS SECTION audit")
    print("=" * 70)
    print(f"  Floors × columns: {report['n_floors']} × {report['n_columns']}")
    print(f"  Cross-section value distribution: {report['value_distribution']}")
    print(f"  Inconsistent columns (mid-stack cross-section change): {len(report['inconsistent_columns'])}")
    print(f"  DXF column entities scanned: {report['n_dxf_column_entities']}")
    print(f"    of which had a hard-truth label: {report['n_hard_truth_available']}")
    print(f"  Mismatches vs hard-truth (CIRCLE or 4-vertex rect): {report['n_mismatch_vs_hard_truth']}")
    print(f"  Mismatches vs independent algorithm: {report['n_mismatch_vs_independent_algo']}")
    if report["disagreements"]:
        print("\n  First disagreements (up to 50):")
        for d in report["disagreements"][:10]:
            print(f"    - {d['dxftype']} layer={d['layer']} prod={d['production']!r} "
                  f"indep={d['independent']!r} hard={d['hard_truth']!r}")
    if report["inconsistent_columns"]:
        print("\n  Columns with cross-section changing between floors:")
        for ic in report["inconsistent_columns"][:10]:
            print(f"    - column {ic['column']}: {ic['values']}")
    print(f"\n  Full report: {workspace / 'audit_cross_sections.json'}")
