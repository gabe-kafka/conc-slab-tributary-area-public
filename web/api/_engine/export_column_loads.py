"""
Excel export module for column load takedown data.

This module exports column tributary area data to Excel format for structural
engineering load takedown calculations in a master matrix format.
"""

import math
import re

import openpyxl
from openpyxl.styles import Alignment, Font
from shapely.geometry import Point

from fascade_utils import (
    compute_fascade_assignments,
    FASCADE_DISTANCE_THRESHOLD,
    FASCADE_SAMPLE_SPACING,
)


RANGE_FLOOR_RE = re.compile(r'^\s*([A-Za-z]*)(\d+)\s*[-–—]\s*([A-Za-z]*)(\d+)\s*$')
UPPER_LAYOUT_FLOOR_BASE = 900000.0

CORNER_ANGLE_THRESHOLD = 160.0  # degrees — vertices with interior angle below this are "true corners"
MIN_ALIGNMENT_LABELS = 3
MAX_ALIGNMENT_RESIDUAL_FEET = 8.0


def alphanumeric_sort_key(label):
    """
    Sort key function for alphanumeric sorting.
    Handles labels like: 1, 2, 10, 23, A1, B2, UNLABELED_0, etc.
    """
    try:
        # Convert label to string if it isn't already
        if isinstance(label, bytes):
            label = label.decode('utf-8', errors='replace')
        elif not isinstance(label, str):
            label = str(label) if label is not None else ""
        
        # Split label into numeric and non-numeric parts
        parts = re.split(r'(\d+)', label)
        # Convert numeric parts to integers for proper sorting
        return [int(part) if part.isdigit() else part.lower() for part in parts]
    except Exception:
        # Fallback to simple string sorting if there's an error
        return [str(label)]


def _is_layout_ordered_upper_floor(floor_id):
    floor_str = str(floor_id).upper()
    compact_floor_str = re.sub(r'[^A-Z0-9]', '', floor_str)
    return (
        'ROOF' in floor_str
        or 'BULKHEAD' in floor_str
        or 'BULK HEAD' in floor_str
        or compact_floor_str.startswith('EMR')
        or 'ELEVATOR MACHINE ROOM' in floor_str
        or 'MACHINE ROOM' in floor_str
        or 'PENTHOUSE' in floor_str
        or floor_str == 'PH'
        or compact_floor_str.startswith('PH')
    )


def floor_sort_key(floor_id, floor_positions=None):
    """
    Sort key function for floor identifiers in descending order.
    Handles: MAIN ROOF, ROOF, PENTHOUSE, numbered floors, GROUND, BASEMENT, etc.
    """
    floor_str = str(floor_id).upper()
    compact_floor_str = re.sub(r'[^A-Z0-9]', '', floor_str)

    if floor_positions and _is_layout_ordered_upper_floor(floor_id):
        layout_x = floor_positions.get(str(floor_id).strip())
        if layout_x is not None:
            return (UPPER_LAYOUT_FLOOR_BASE + float(layout_x), floor_str)
    
    # Define priority order (higher number = appears first)
    if (
        compact_floor_str.startswith('EMR')
        or 'ELEVATOR MACHINE ROOM' in floor_str
        or 'MACHINE ROOM' in floor_str
    ):
        match = re.search(r'(\d+)', compact_floor_str)
        return (970 + (int(match.group(1)) if match else 0), floor_str)
    elif 'BULKHEAD' in floor_str or 'BULK HEAD' in floor_str:
        return (950, floor_str)
    elif 'ROOF' in floor_str and 'MAIN' in floor_str:
        return (1000, floor_str)
    elif 'ROOF' in floor_str:
        return (900, floor_str)
    elif 'PENTHOUSE' in floor_str or 'PH' in floor_str:
        return (800, floor_str)
    elif 'GROUND' in floor_str or 'MAIN' in floor_str or 'LOBBY' in floor_str:
        return (-100, floor_str)
    elif 'BASEMENT' in floor_str or floor_str.startswith('B'):
        # Extract basement number if present
        match = re.search(r'B?(\d+)', floor_str)
        if match:
            basement_num = int(match.group(1))
            return (-200 - basement_num, floor_str)
        return (-200, floor_str)
    else:
        # Try to extract numeric floor number
        match = re.search(r'(\d+)', floor_str)
        if match:
            floor_num = int(match.group(1))
            return (floor_num, floor_str)
        # Default: alphabetical
        return (0, floor_str)


def expand_floor_identifier(floor_id):
    """
    Expand grouped floor labels such as ``4-8`` or ``B1-B3`` into
    individual floor identifiers for spreadsheet output.
    """
    floor_str = str(floor_id).strip()
    match = RANGE_FLOOR_RE.fullmatch(floor_str)
    if not match:
        return [floor_str]

    start_prefix, start_num, end_prefix, end_num = match.groups()
    if start_prefix.upper() != end_prefix.upper():
        return [floor_str]

    start_value = int(start_num)
    end_value = int(end_num)
    step = 1 if end_value >= start_value else -1
    prefix = start_prefix.upper()
    return [f"{prefix}{value}" if prefix else str(value) for value in range(start_value, end_value + step, step)]


def expand_floor_map(source_floor_data):
    """
    Duplicate grouped floor data into one spreadsheet row per represented floor.
    """
    expanded = {}
    for floor_id, value_map in source_floor_data.items():
        expanded_floor_ids = expand_floor_identifier(floor_id)
        for expanded_floor_id in expanded_floor_ids:
            if expanded_floor_id not in expanded:
                expanded[expanded_floor_id] = {}
            expanded[expanded_floor_id].update(value_map)
    return expanded


def expand_floor_scalar_map(source_values):
    """Duplicate per-floor scalar values for grouped floor identifiers."""
    expanded = {}
    for floor_id, value in source_values.items():
        for expanded_floor_id in expand_floor_identifier(floor_id):
            expanded[expanded_floor_id] = value
    return expanded


def _floor_plan_layout_x(floor_plan):
    datum_info = floor_plan.get('alignment_datum') or {}
    if isinstance(datum_info, dict):
        datum = datum_info.get('point')
        if datum is not None:
            return float(datum[0])

    user_datum = floor_plan.get('user_datum')
    if user_datum is not None:
        return float(user_datum[0])

    slab = floor_plan.get('slab_polygon')
    if slab is not None and not slab.is_empty:
        min_x, _, max_x, _ = slab.bounds
        return (float(min_x) + float(max_x)) / 2.0

    return None


def build_floor_position_overrides(floor_plans):
    """Return x-position sort overrides for upper non-numeric floor labels."""
    positions = {}
    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        if not _is_layout_ordered_upper_floor(floor_id):
            continue

        layout_x = _floor_plan_layout_x(floor_plan)
        if layout_x is None:
            continue

        for expanded_floor_id in expand_floor_identifier(floor_id):
            positions[str(expanded_floor_id).strip()] = layout_x

    return positions


def sort_floor_ids(floor_ids, floor_positions=None):
    return sorted(
        list(floor_ids),
        key=lambda floor_id: floor_sort_key(floor_id, floor_positions),
        reverse=True,
    )


def _vertex_angle(prev, vertex, nxt):
    """Compute interior angle at a polygon vertex in degrees."""
    v1 = (prev[0] - vertex[0], prev[1] - vertex[1])
    v2 = (nxt[0] - vertex[0], nxt[1] - vertex[1])
    mag1 = math.hypot(v1[0], v1[1])
    mag2 = math.hypot(v2[0], v2[1])
    if mag1 < 1e-10 or mag2 < 1e-10:
        return 180.0
    cos_angle = max(-1.0, min(1.0, (v1[0] * v2[0] + v1[1] * v2[1]) / (mag1 * mag2)))
    return math.degrees(math.acos(cos_angle))


def _extract_building_corners(slab_polygon):
    """Extract true building corners using the convex hull.

    The convex hull smooths out notches, setbacks, and re-entrant corners,
    leaving only the vertices that define the overall building footprint.
    Vertices where the hull turns significantly (< CORNER_ANGLE_THRESHOLD)
    are classified as building corners.
    """
    hull = slab_polygon.convex_hull
    if hull.geom_type != 'Polygon':
        return []

    coords = list(hull.exterior.coords)
    if coords[0] == coords[-1]:
        coords = coords[:-1]
    n = len(coords)
    if n < 3:
        return []

    corners = []
    for i in range(n):
        prev = coords[(i - 1) % n]
        curr = coords[i]
        nxt = coords[(i + 1) % n]
        angle = _vertex_angle(prev, curr, nxt)
        if angle < CORNER_ANGLE_THRESHOLD:
            corners.append(Point(curr))
    return corners


def _compute_edge_threshold(column_points, slab_polygon):
    """Compute a threshold separating edge columns from interior columns.

    Uses the largest gap in sorted column-to-boundary distances.
    Falls back to half the median distance when no clear gap exists.
    """
    if not column_points:
        return 5.0
    boundary = slab_polygon.boundary
    dists = sorted(pt.distance(boundary) for pt in column_points)
    if len(dists) < 2:
        return dists[0] + 1.0

    max_gap = 0.0
    threshold = dists[-1] / 2.0
    for i in range(len(dists) - 1):
        gap = dists[i + 1] - dists[i]
        if gap > max_gap:
            max_gap = gap
            threshold = (dists[i] + dists[i + 1]) / 2.0
    return threshold


def classify_columns_for_floor(floor_plan):
    """Classify all columns on a floor as center/edge/corner.

    Returns list of KLL values parallel to column_points:
      4 = center
      3 = edge
      2 = corner
      None = no slab at this column on this floor (rescued orphan in an
             opening) — KLL doesn't apply because no live load is being
             reduced at this level for this column.

    Edge threshold is computed from columns *inside* the slab only, so
    orphans far outside don't skew the gap detection.
    """
    column_points = floor_plan.get('column_points', [])
    slab_polygon = floor_plan.get('slab_polygon')

    if not column_points or slab_polygon is None:
        return [4] * len(column_points)

    boundary = slab_polygon.boundary
    true_corners = _extract_building_corners(slab_polygon)

    # Inside-slab columns only contribute to the edge-threshold sample.
    inside_pts = [pt for pt in column_points if slab_polygon.covers(pt)]
    threshold = _compute_edge_threshold(inside_pts or column_points, slab_polygon)

    kll_values = []
    for col_pt in column_points:
        if not slab_polygon.covers(col_pt):
            # Column sits in an opening / outside the slab — no slab attached
            # at this floor, so KLL is undefined for this level.
            kll_values.append(None)
            continue

        dist_to_edge = col_pt.distance(boundary)

        if dist_to_edge >= threshold:
            kll_values.append(4)  # center
            continue

        # Edge column — check if near a true corner
        if true_corners:
            min_corner_dist = min(col_pt.distance(c) for c in true_corners)
            if min_corner_dist <= threshold:
                kll_values.append(2)  # corner
                continue

        kll_values.append(3)  # edge

    return kll_values


CIRCLE_CIRCULARITY_THRESHOLD = 0.96   # raised from 0.92 to exclude octagons (≈0.948)
OCTAGON_CIRCULARITY_LOWER = 0.92      # 8-vertex polygons above this read as octagons
NON_CONVEX_AREA_RATIO = 0.85          # poly.area / hull.area below → L/T/+ shape
SHAPE_TAG_BY_VERTEX = {3: "tri", 5: "pent", 6: "hex"}


def _round_in(val_in: float) -> int:
    """Half-away-from-zero rounding to nearest inch."""
    return int(math.floor(val_in + 0.5)) if val_in >= 0 else -int(math.floor(-val_in + 0.5))


def _mrr_sides_in(footprint) -> tuple[int, int] | None:
    """Return (w, d) in inches from the minimum-rotated-rectangle, w ≤ d."""
    min_rect = footprint.minimum_rotated_rectangle
    coords = list(min_rect.exterior.coords)
    if len(coords) < 5:
        return None
    side_a = math.hypot(coords[1][0] - coords[0][0], coords[1][1] - coords[0][1]) * 12.0
    side_b = math.hypot(coords[2][0] - coords[1][0], coords[2][1] - coords[1][1]) * 12.0
    return _round_in(min(side_a, side_b)), _round_in(max(side_a, side_b))


def _bbox_sides_in(footprint) -> tuple[int, int]:
    """Return (w, d) of the axis-aligned bounding box in inches, w ≤ d."""
    min_x, min_y, max_x, max_y = footprint.bounds
    w = _round_in((max_x - min_x) * 12.0)
    h = _round_in((max_y - min_y) * 12.0)
    return min(w, h), max(w, h)


def format_cross_section(footprint):
    """
    Format a column footprint polygon as a cross-section label.

    Outputs (all dims in inches):
      - "12x24"      rectangle / rotated rectangle / diamond (MRR sides)
      - "d24"        circle (high circularity, dia from area)
      - "oct24"      octagon (8 vertices, circularity 0.92–0.96)
      - "tri18x24"   triangle (3-vertex polygon, bbox dims)
      - "pent18x24"  pentagon (5-vertex)
      - "hex18x24"   hexagon (6-vertex)
      - "L24x26"     non-convex (L / T / + / U)
      - None if footprint missing or degenerate
    """
    if footprint is None or footprint.is_empty:
        return None

    area_ft2 = footprint.area
    perimeter_ft = footprint.length
    if area_ft2 <= 1e-9 or perimeter_ft <= 1e-9:
        return None

    coords = list(footprint.exterior.coords)
    if coords and coords[0] == coords[-1]:
        coords = coords[:-1]
    n = len(coords)

    circularity = 4.0 * math.pi * area_ft2 / (perimeter_ft * perimeter_ft)

    # True circle (raised threshold so octagons fall through).
    if circularity >= CIRCLE_CIRCULARITY_THRESHOLD:
        diameter_in = 2.0 * math.sqrt(area_ft2 / math.pi) * 12.0
        return f"d{_round_in(diameter_in)}"

    # Octagon: 8 vertices in the threshold band.
    if n == 8 and circularity >= OCTAGON_CIRCULARITY_LOWER:
        diameter_in = 2.0 * math.sqrt(area_ft2 / math.pi) * 12.0
        return f"oct{_round_in(diameter_in)}"

    # Low-vertex non-rectangular polygons (triangle, pentagon, hexagon).
    # Use MRR so the dim reflects the column's actual oriented extent,
    # not the drawing-orientation bbox.
    if n in SHAPE_TAG_BY_VERTEX:
        mrr = _mrr_sides_in(footprint)
        if mrr is not None:
            w, d = mrr
            return f"{SHAPE_TAG_BY_VERTEX[n]}{w}x{d}"

    # Non-convex (L / T / + / U): convex-hull area materially larger than polygon.
    try:
        hull_area = footprint.convex_hull.area
        if hull_area > 1e-9 and area_ft2 / hull_area < NON_CONVEX_AREA_RATIO:
            mrr = _mrr_sides_in(footprint)
            if mrr is not None:
                w, d = mrr
                return f"L{w}x{d}"
    except Exception:
        pass

    # Default: rectangle / rotated rectangle / diamond — MRR sides.
    mrr = _mrr_sides_in(footprint)
    if mrr is None:
        return None
    w, d = mrr
    return f"{w}x{d}"


def collect_cross_section_data(floor_plans):
    """
    Collect column cross-section dimensions per floor in the same matrix shape
    as collect_master_matrix_data. Cell values are formatted strings (e.g.
    "12x24", "24x24", "d24") or None when the column is absent on a floor.
    """
    all_column_labels = set()
    floor_data = {}

    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        column_points = floor_plan.get('column_points', [])
        column_labels = floor_plan.get('column_labels', [])
        column_footprints = floor_plan.get('column_footprints', [])
        point_metadata = floor_plan.get('point_metadata', [])

        if floor_id not in floor_data:
            floor_data[floor_id] = {}

        if point_metadata:
            for metadata in point_metadata:
                if metadata.get('type') != 'column':
                    continue
                col_idx = metadata.get('column_index')
                if col_idx is None or col_idx >= len(column_points):
                    continue

                if col_idx < len(column_labels):
                    label = column_labels[col_idx]
                    if not isinstance(label, str):
                        label = str(label) if label is not None else f"UNLABELED_{col_idx}"
                else:
                    label = f"UNLABELED_{col_idx}"

                footprint = column_footprints[col_idx] if col_idx < len(column_footprints) else None
                floor_data[floor_id][label] = format_cross_section(footprint)
                all_column_labels.add(label)
        else:
            for col_idx, _point in enumerate(column_points):
                label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
                if not isinstance(label, str):
                    label = str(label) if label is not None else f"UNLABELED_{col_idx}"

                footprint = column_footprints[col_idx] if col_idx < len(column_footprints) else None
                floor_data[floor_id][label] = format_cross_section(footprint)
                all_column_labels.add(label)

    floor_data = expand_floor_map(floor_data)

    sorted_column_labels = sorted(list(all_column_labels), key=alphanumeric_sort_key)
    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floor_numbers = sort_floor_ids(floor_data.keys(), floor_positions)

    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for column_label in sorted_column_labels:
            matrix[floor_id][column_label] = floor_data[floor_id].get(column_label)

    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_column_labels,
        'matrix': matrix,
    }


def _floor_alignment_origin(floor_plan):
    """Per-floor datum point for aligning column footprints across floors.

    Prefers the length-weighted centroid of wall lines (core/stair walls
    are typically the most stable feature floor-to-floor). Falls back to
    the slab polygon centroid when no walls are present.
    """
    from shapely.ops import unary_union

    walls = floor_plan.get("walls") or []
    wall_lines = [
        w.get("wall_line") for w in walls
        if w.get("wall_line") is not None and not w["wall_line"].is_empty
    ]
    if wall_lines:
        merged = unary_union(wall_lines)
        if not merged.is_empty:
            return merged.centroid

    slab = floor_plan.get("slab_polygon")
    if slab is not None and not slab.is_empty:
        return slab.centroid

    return None


def compute_floor_datums(floor_plans):
    """
    Per-floor alignment datum used for cross-floor geometric comparison.

    Convention: floor plans are drawn stacked in DXF model space at a
    fixed vertical offset (3000 inches per floor for the demo DXF). The
    slab bbox min corner naturally encodes whatever stacking offset the
    user used — translations between adjacent floors come out to exactly
    the stacking offset without hardcoding it.

    AI selection (datum_utils.select_alignment_datums) and wall/slab
    centroid remain available as fallbacks for non-stacked layouts.

    Returns {floor_id: {"point": (x, y), "source": str}}.
    """
    result = {}
    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        if floor_id in result:
            continue

        user_datum = floor_plan.get('user_datum')
        if user_datum is not None:
            result[floor_id] = {"point": (user_datum[0], user_datum[1]), "source": "user"}
            continue

        slab = floor_plan.get('slab_polygon')
        if slab is not None and not slab.is_empty:
            min_x, min_y, _, _ = slab.bounds
            result[floor_id] = {"point": (min_x, min_y), "source": "stacked"}
            continue

        wall_origin = _floor_alignment_origin(floor_plan)
        if wall_origin is None:
            result[floor_id] = {"point": None, "source": "none"}
            continue

        has_walls = any(
            w.get("wall_line") is not None and not w["wall_line"].is_empty
            for w in (floor_plan.get("walls") or [])
        )
        source = "wall_centroid" if has_walls else "slab_centroid"
        result[floor_id] = {"point": (wall_origin.x, wall_origin.y), "source": source}

    return result


def _normalize_alignment_label(label):
    value = re.sub(r'\s+', '', str(label or '').strip().upper())
    if not value or 'UNLABELED' in value:
        return None
    return value


def _median(values):
    values = list(values)
    if not values:
        return 0.0
    sorted_values = sorted(values)
    mid = len(sorted_values) // 2
    if len(sorted_values) % 2 == 1:
        return sorted_values[mid]
    return (sorted_values[mid - 1] + sorted_values[mid]) / 2.0


def _column_label_centers(columns):
    grouped = {}
    for label, point, _footprint in columns:
        normalized = _normalize_alignment_label(label)
        if not normalized or point is None:
            continue
        current = grouped.setdefault(normalized, [0.0, 0.0, 0])
        current[0] += point.x
        current[1] += point.y
        current[2] += 1

    return {
        label: (total_x / count, total_y / count)
        for label, (total_x, total_y, count) in grouped.items()
        if count > 0
    }


def _median_label_offset(upper_columns, lower_columns):
    upper_centers = _column_label_centers(upper_columns)
    lower_centers = _column_label_centers(lower_columns)
    dxs = []
    dys = []

    for label, upper_point in upper_centers.items():
        lower_point = lower_centers.get(label)
        if lower_point is None:
            continue
        dxs.append(lower_point[0] - upper_point[0])
        dys.append(lower_point[1] - upper_point[1])

    if len(dxs) < MIN_ALIGNMENT_LABELS:
        return None

    dx = _median(dxs)
    dy = _median(dys)
    residual = _median(
        math.hypot(candidate_dx - dx, dys[index] - dy)
        for index, candidate_dx in enumerate(dxs)
    )

    if residual > MAX_ALIGNMENT_RESIDUAL_FEET:
        return None

    return dx, dy


def collect_column_discontinuities(floor_plans, floor_datums=None):
    """
    Polygon-overlap discontinuity check, with floors aligned via the
    per-floor datums from compute_floor_datums (or computed inline when
    not provided).

    A column on floor F is continuous iff its (translated) footprint
    intersects any column footprint on the floor immediately below.

    Bottom floor has no discontinuities — those columns are assumed to
    continue to foundation.

    Returns: { floor_id: set((label, point_x, point_y)) }
    """
    from shapely.affinity import translate
    from shapely.geometry import Point

    if floor_datums is None:
        floor_datums = compute_floor_datums(floor_plans)

    floor_data = {}  # floor_id -> {origin, source, columns: [(label, point, footprint)]}
    floor_order = []

    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        column_labels = floor_plan.get('column_labels', [])
        column_points = floor_plan.get('column_points', [])
        column_footprints = floor_plan.get('column_footprints', [])

        if floor_id not in floor_data:
            datum_info = floor_datums.get(floor_id, {})
            datum = datum_info.get("point")
            origin = Point(datum[0], datum[1]) if datum is not None else None
            floor_data[floor_id] = {
                "origin": origin,
                "source": datum_info.get("source", "none"),
                "columns": [],
            }
            floor_order.append(floor_id)

        for col_idx in range(len(column_labels)):
            label = column_labels[col_idx]
            if not isinstance(label, str):
                label = str(label) if label is not None else f"UNLABELED_{col_idx}"
            footprint = column_footprints[col_idx] if col_idx < len(column_footprints) else None
            point = column_points[col_idx] if col_idx < len(column_points) else None
            if footprint is None or footprint.is_empty or point is None:
                continue
            floor_data[floor_id]["columns"].append((label, point, footprint))

    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floors = [
        floor_id for floor_id in sort_floor_ids(floor_order, floor_positions)
        if floor_data[floor_id]["columns"]
    ]

    discontinuities = {floor_id: set() for floor_id in floor_order}
    for upper, lower in zip(sorted_floors[:-1], sorted_floors[1:]):
        upper_d = floor_data[upper]
        lower_d = floor_data[lower]
        label_offset = _median_label_offset(upper_d["columns"], lower_d["columns"])
        if label_offset is not None:
            dx, dy = label_offset
        elif upper_d["origin"] is None or lower_d["origin"] is None:
            # No datum — fall back to label match for this pair.
            lower_labels = {label for label, _, _ in lower_d["columns"]}
            for label, point, _ in upper_d["columns"]:
                if label not in lower_labels:
                    discontinuities[upper].add((label, round(point.x, 2), round(point.y, 2)))
            continue
        else:
            dx = lower_d["origin"].x - upper_d["origin"].x
            dy = lower_d["origin"].y - upper_d["origin"].y

        lower_footprints = [fp for _, _, fp in lower_d["columns"]]

        # 1-foot proximity tolerance: a column is continuous if its
        # (datum-translated) footprint comes within 1 ft of any lower
        # footprint. Catches near-misses from drafting drift, shifted
        # column sizes, or slight datum offsets.
        proximity_tol_ft = 1.0
        for label, point, upper_fp in upper_d["columns"]:
            translated = translate(upper_fp, xoff=dx, yoff=dy)
            if not any(translated.distance(lp) <= proximity_tol_ft for lp in lower_footprints):
                discontinuities[upper].add((label, round(point.x, 2), round(point.y, 2)))

    return discontinuities


def collect_master_matrix_data(floor_plans):
    """
    Collect and organize data for master matrix format.
    
    Args:
        floor_plans: List of floor plan dictionaries with column data
        
    Returns:
        dict with keys:
            'floor_numbers': List of floor identifiers (sorted descending)
            'column_labels': List of unique column labels (sorted ascending)
            'matrix': Dict[floor][column] = tributary_area (rounded up)
    """
    # Collect all unique column labels across all floor plans
    all_column_labels = set()
    
    # Build a mapping: floor_id -> {column_label -> tributary_area}
    floor_data = {}
    
    for floor_plan in floor_plans:
        # Get floor identifier (use floor_number if available, otherwise boundary_id)
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        
        column_points = floor_plan.get('column_points', [])
        column_labels = floor_plan.get('column_labels', [])
        areas = floor_plan.get('areas', [])
        point_metadata = floor_plan.get('point_metadata', [])
        
        # Initialize floor data
        if floor_id not in floor_data:
            floor_data[floor_id] = {}
        
        if point_metadata:
            # Legacy mixed-site shape: iterate through metadata to extract column tributary areas.
            for i, metadata in enumerate(point_metadata):
                if metadata.get('type') != 'column':
                    continue

                col_idx = metadata.get('column_index')
                if col_idx is None or col_idx >= len(column_points):
                    continue

                if col_idx < len(column_labels):
                    label = column_labels[col_idx]
                    if not isinstance(label, str):
                        label = str(label) if label is not None else f"UNLABELED_{col_idx}"
                else:
                    label = f"UNLABELED_{col_idx}"

                tributary_area = areas[i] if i < len(areas) else 0.0
                rounded_area = math.ceil(tributary_area)
                floor_data[floor_id][label] = rounded_area
                all_column_labels.add(label)
        else:
            # Column-only shape: `areas` already aligns with `column_points` / `column_labels`.
            for col_idx, point in enumerate(column_points):
                label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
                if not isinstance(label, str):
                    label = str(label) if label is not None else f"UNLABELED_{col_idx}"

                tributary_area = areas[col_idx] if col_idx < len(areas) else 0.0
                rounded_area = math.ceil(tributary_area)
                floor_data[floor_id][label] = rounded_area
                all_column_labels.add(label)
    
    floor_data = expand_floor_map(floor_data)

    # Sort column labels alphanumerically (ascending: 1, 2, 3, ..., N)
    sorted_column_labels = sorted(list(all_column_labels), key=alphanumeric_sort_key)
    
    # Sort floor numbers in descending order (MAIN ROOF → 2ND → 1ST)
    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floor_numbers = sort_floor_ids(floor_data.keys(), floor_positions)
    
    # Build matrix data structure: matrix[floor][column] = tributary_area
    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for column_label in sorted_column_labels:
            # Get tributary area if column exists on this floor, otherwise None
            matrix[floor_id][column_label] = floor_data[floor_id].get(column_label, None)
    
    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_column_labels,
        'matrix': matrix
    }


def collect_additional_load_matrix_data(floor_plans):
    """
    Per-column, per-floor SF coming from non-BOUNDARY load zones
    (Terrace, Balcony, etc.) — i.e. the portion of each Voronoi region
    that lies under a secondary load layer. Blank = column not on floor.
    """
    all_column_labels = set()
    floor_data = {}

    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        column_labels = floor_plan.get('column_labels', [])
        column_points = floor_plan.get('column_points', [])
        column_load_areas = floor_plan.get('column_load_areas', [])

        if floor_id not in floor_data:
            floor_data[floor_id] = {}

        for col_idx, _point in enumerate(column_points):
            label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
            if not isinstance(label, str):
                label = str(label) if label is not None else f"UNLABELED_{col_idx}"

            load_areas = column_load_areas[col_idx] if col_idx < len(column_load_areas) else []
            secondary_area = 0.0
            for item in load_areas:
                layer = str(item.get('layer') or 'BOUNDARY').upper()
                if layer == 'BOUNDARY':
                    continue
                secondary_area += float(item.get('area', 0.0) or 0.0)

            floor_data[floor_id][label] = math.ceil(secondary_area)
            all_column_labels.add(label)

    floor_data = expand_floor_map(floor_data)

    sorted_column_labels = sorted(list(all_column_labels), key=alphanumeric_sort_key)
    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floor_numbers = sort_floor_ids(floor_data.keys(), floor_positions)

    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for column_label in sorted_column_labels:
            matrix[floor_id][column_label] = floor_data[floor_id].get(column_label, None)

    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_column_labels,
        'matrix': matrix,
    }


def collect_load_zone_area_rows(floor_plans):
    """Return long-form tributary area rows split by boundary/load layer."""
    rows = []

    for floor_plan in floor_plans:
        source_floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        expanded_floor_ids = expand_floor_identifier(source_floor_id)
        column_labels = floor_plan.get('column_labels', [])
        column_load_areas = floor_plan.get('column_load_areas', [])

        for col_idx, load_areas in enumerate(column_load_areas):
            label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
            if not isinstance(label, str):
                label = str(label) if label is not None else f"UNLABELED_{col_idx}"

            for item in load_areas:
                area = float(item.get('area', 0.0) or 0.0)
                if area <= 1e-6:
                    continue

                for floor_id in expanded_floor_ids:
                    rows.append({
                        'floor': floor_id,
                        'load_zone': item.get('layer') or 'BOUNDARY',
                        'column': label,
                        'area': area,
                    })

    floor_positions = build_floor_position_overrides(floor_plans)
    return sorted(
        rows,
        key=lambda row: (
            -floor_sort_key(row['floor'], floor_positions)[0],
            str(row['floor']),
            str(row['load_zone']),
            alphanumeric_sort_key(row['column']),
        ),
    )


def collect_kll_matrix_data(floor_plans):
    """
    Collect KLL (live load element factor) values for each column on each floor.

    Classification: center=4, edge=3, corner=2 based on proximity to slab edge.

    Returns dict with same structure as collect_master_matrix_data:
        'floor_numbers', 'column_labels', 'matrix' (values are KLL integers).
    """
    all_column_labels = set()
    floor_data = {}

    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        column_points = floor_plan.get('column_points', [])
        column_labels = floor_plan.get('column_labels', [])
        point_metadata = floor_plan.get('point_metadata', [])

        kll_values = classify_columns_for_floor(floor_plan)

        if floor_id not in floor_data:
            floor_data[floor_id] = {}

        if point_metadata:
            for i, metadata in enumerate(point_metadata):
                if metadata.get('type') != 'column':
                    continue

                col_idx = metadata.get('column_index')
                if col_idx is None or col_idx >= len(column_points):
                    continue

                if col_idx < len(column_labels):
                    label = column_labels[col_idx]
                    if not isinstance(label, str):
                        label = str(label) if label is not None else f"UNLABELED_{col_idx}"
                else:
                    label = f"UNLABELED_{col_idx}"

                floor_data[floor_id][label] = kll_values[col_idx] if col_idx < len(kll_values) else 4
                all_column_labels.add(label)
        else:
            for col_idx, point in enumerate(column_points):
                label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
                if not isinstance(label, str):
                    label = str(label) if label is not None else f"UNLABELED_{col_idx}"

                floor_data[floor_id][label] = kll_values[col_idx] if col_idx < len(kll_values) else 4
                all_column_labels.add(label)

    floor_data = expand_floor_map(floor_data)

    sorted_column_labels = sorted(list(all_column_labels), key=alphanumeric_sort_key)
    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floor_numbers = sort_floor_ids(floor_data.keys(), floor_positions)

    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for column_label in sorted_column_labels:
            matrix[floor_id][column_label] = floor_data[floor_id].get(column_label, None)

    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_column_labels,
        'matrix': matrix
    }


def collect_fascade_length_data(
    floor_plans,
    distance_threshold=FASCADE_DISTANCE_THRESHOLD,
    sample_spacing=FASCADE_SAMPLE_SPACING
):
    """Aggregate façade length attribution data for every floor."""
    floor_data = {}
    all_labels = set()
    floor_perimeters = {}
    floor_assigned_totals = {}
    floor_thresholds = {}
    floor_coverages = {}
    floor_max_distances = {}
    floor_methods = {}
    
    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        if floor_plan.get('fascade_data'):
            result = floor_plan['fascade_data']
        else:
            result = compute_fascade_assignments(
                floor_plan,
                distance_threshold=distance_threshold,
                sample_spacing=sample_spacing
            )
            floor_plan['fascade_data'] = result
        
        length_map = result.get('length_map', {})
        floor_perimeters[floor_id] = result.get('perimeter', 0.0)
        floor_assigned_totals[floor_id] = result.get('assigned_total', 0.0)
        floor_thresholds[floor_id] = result.get('threshold_used', distance_threshold)
        floor_coverages[floor_id] = result.get('coverage_ratio', 0.0)
        floor_max_distances[floor_id] = result.get('max_distance_seen', 0.0)
        floor_methods[floor_id] = result.get('assignment_method', 'nearest_boundary_sample')
        
        if floor_id not in floor_data:
            floor_data[floor_id] = {}
        
        for label, length in length_map.items():
            floor_data[floor_id][label] = floor_data[floor_id].get(label, 0.0) + length
            all_labels.add(label)
    
    floor_data = expand_floor_map(floor_data)
    floor_perimeters = expand_floor_scalar_map(floor_perimeters)
    floor_assigned_totals = expand_floor_scalar_map(floor_assigned_totals)
    floor_thresholds = expand_floor_scalar_map(floor_thresholds)
    floor_coverages = expand_floor_scalar_map(floor_coverages)
    floor_max_distances = expand_floor_scalar_map(floor_max_distances)
    floor_methods = expand_floor_scalar_map(floor_methods)

    floor_positions = build_floor_position_overrides(floor_plans)
    sorted_floor_numbers = sort_floor_ids(floor_data.keys(), floor_positions)
    sorted_labels = sorted(list(all_labels), key=alphanumeric_sort_key)
    
    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for label in sorted_labels:
            matrix[floor_id][label] = floor_data.get(floor_id, {}).get(label)
    
    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_labels,
        'matrix': matrix,
        'perimeters': floor_perimeters,
        'assigned_totals': floor_assigned_totals,
        'thresholds': floor_thresholds,
        'coverage': floor_coverages,
        'max_distances': floor_max_distances,
        'methods': floor_methods,
    }


def export_column_load_takedown(floor_plans, output_filename="column_load_takedown.xlsx"):
    """
    Export column load takedown data to Excel in master matrix format.
    
    Args:
        floor_plans: List of floor plan dictionaries with column data
        output_filename: Path to output Excel file
        
    Returns:
        None (writes file to disk)
    """
    try:
        # Collect master matrix data
        matrix_data = collect_master_matrix_data(floor_plans)
        
        floor_numbers = matrix_data['floor_numbers']
        column_labels = matrix_data['column_labels']
        matrix = matrix_data['matrix']
        
        # Create new Workbook object
        workbook = openpyxl.Workbook()
        
        # Remove the default sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create single worksheet named "MASTER TRIBUTARY AREA"
        worksheet = workbook.create_sheet(title="MASTER TRIBUTARY AREA")
        
        # Write header row: "Slab #" in first column, then column labels
        header_row = ["Slab #"] + column_labels
        worksheet.append(header_row)
        
        # Apply bold formatting to header row
        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font
        
        # Write floor rows
        for floor_id in floor_numbers:
            # Start with floor identifier in first column
            row_data = [floor_id]
            
            # Add tributary areas for each column
            for column_label in column_labels:
                area = matrix[floor_id].get(column_label)
                # Leave cell empty if column doesn't exist on this floor
                row_data.append(area if area is not None else "")
            
            worksheet.append(row_data)
        
        # Apply formatting
        # Bold first column (floor numbers)
        for row in worksheet.iter_rows(min_row=2, max_row=len(floor_numbers) + 1, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)
        
        # Right-align numeric cells (all columns except first)
        right_align = Alignment(horizontal='right')
        for row in worksheet.iter_rows(min_row=2, max_row=len(floor_numbers) + 1, min_col=2):
            for cell in row:
                if cell.value and cell.value != "":
                    cell.alignment = right_align
        
        # Set column widths
        # First column (Floor) - wider for floor names
        worksheet.column_dimensions['A'].width = 15
        
        # Other columns - auto-size based on content
        for col_idx, column_label in enumerate(column_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            # Calculate width based on label length and typical area values
            label_width = len(str(column_label))
            worksheet.column_dimensions[col_letter].width = max(label_width + 2, 8)
        
        # Freeze top row and first column for scrolling
        worksheet.freeze_panes = 'B2'

        # --- Master Additional Load Area Sheet ---
        add_data = collect_additional_load_matrix_data(floor_plans)
        add_labels = add_data['column_labels']
        add_floor_numbers = add_data['floor_numbers']
        add_matrix = add_data['matrix']

        add_sheet = workbook.create_sheet(title="MASTER ADDITIONAL LOAD AREA")
        add_sheet.append(["Slab #"] + add_labels)
        for cell in add_sheet[1]:
            cell.font = header_font

        for floor_id in add_floor_numbers:
            row_data = [floor_id]
            for label in add_labels:
                value = add_matrix[floor_id].get(label)
                row_data.append(value if value is not None else "")
            add_sheet.append(row_data)

        if add_floor_numbers:
            for row in add_sheet.iter_rows(
                min_row=2, max_row=len(add_floor_numbers) + 1,
                min_col=1, max_col=1
            ):
                for cell in row:
                    cell.font = Font(bold=True)

        for row in add_sheet.iter_rows(
            min_row=2, max_row=len(add_floor_numbers) + 1, min_col=2
        ):
            for cell in row:
                if cell.value not in ("", None):
                    cell.alignment = right_align

        add_sheet.column_dimensions['A'].width = 15
        for col_idx, label in enumerate(add_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            label_width = len(str(label))
            add_sheet.column_dimensions[col_letter].width = max(label_width + 2, 8)

        add_sheet.freeze_panes = 'B2'

        # --- Master Cross Section Sheet ---
        xs_data = collect_cross_section_data(floor_plans)
        xs_labels = xs_data['column_labels']
        xs_floor_numbers = xs_data['floor_numbers']
        xs_matrix = xs_data['matrix']

        xs_sheet = workbook.create_sheet(title="MASTER CROSS SECTION")
        xs_sheet.append(["Slab #"] + xs_labels)
        for cell in xs_sheet[1]:
            cell.font = header_font

        for floor_id in xs_floor_numbers:
            row_data = [floor_id]
            for label in xs_labels:
                value = xs_matrix[floor_id].get(label)
                row_data.append(value if value is not None else "")
            xs_sheet.append(row_data)

        if xs_floor_numbers:
            for row in xs_sheet.iter_rows(
                min_row=2, max_row=len(xs_floor_numbers) + 1,
                min_col=1, max_col=1
            ):
                for cell in row:
                    cell.font = Font(bold=True)

        for row in xs_sheet.iter_rows(
            min_row=2, max_row=len(xs_floor_numbers) + 1, min_col=2
        ):
            for cell in row:
                if cell.value not in ("", None):
                    cell.alignment = right_align

        xs_sheet.column_dimensions['A'].width = 15
        for col_idx, label in enumerate(xs_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            label_width = len(str(label))
            xs_sheet.column_dimensions[col_letter].width = max(label_width + 2, 10)

        xs_sheet.freeze_panes = 'B2'

        # --- Load Zone Area Sheet ---
        load_zone_rows = collect_load_zone_area_rows(floor_plans)
        if load_zone_rows:
            load_zone_sheet = workbook.create_sheet(title="LOAD ZONE AREAS")
            load_zone_sheet.append(["Floor", "Load Zone", "Column", "Area (SF)", "Area Rounded (SF)"])
            for cell in load_zone_sheet[1]:
                cell.font = header_font

            for row in load_zone_rows:
                load_zone_sheet.append([
                    row['floor'],
                    row['load_zone'],
                    row['column'],
                    round(row['area'], 2),
                    int(math.ceil(row['area'])),
                ])

            for row in load_zone_sheet.iter_rows(
                min_row=2,
                max_row=len(load_zone_rows) + 1,
                min_col=4,
            ):
                for cell in row:
                    if cell.value not in ("", None):
                        cell.alignment = right_align

            load_zone_sheet.column_dimensions['A'].width = 15
            load_zone_sheet.column_dimensions['B'].width = 24
            load_zone_sheet.column_dimensions['C'].width = 14
            load_zone_sheet.column_dimensions['D'].width = 14
            load_zone_sheet.column_dimensions['E'].width = 18
            load_zone_sheet.freeze_panes = 'A2'

        # --- Fascade Length Sheet ---
        fascade_data = collect_fascade_length_data(floor_plans)
        fascade_labels = fascade_data['column_labels']
        fascade_floor_numbers = fascade_data['floor_numbers']
        fascade_matrix = fascade_data['matrix']
        fascade_perimeters = fascade_data.get('perimeters', {})
        fascade_assigned_totals = fascade_data.get('assigned_totals', {})
        fascade_thresholds = fascade_data.get('thresholds', {})
        fascade_coverage = fascade_data.get('coverage', {})
        fascade_max_distances = fascade_data.get('max_distances', {})
        fascade_methods = fascade_data.get('methods', {})
        
        fascade_sheet = workbook.create_sheet(title="FASCADE LENGTH")
        fascade_header = ["Floor"] + fascade_labels + ["TOTAL ASSIGNED (FT)", "PERIMETER (FT)"]
        fascade_sheet.append(fascade_header)
        for cell in fascade_sheet[1]:
            cell.font = header_font
        
        for floor_id in fascade_floor_numbers:
            row = [floor_id]
            floor_lengths = fascade_matrix.get(floor_id, {})
            for label in fascade_labels:
                length_value = floor_lengths.get(label)
                if length_value is None:
                    row.append("")
                else:
                    rounded_value = int(math.ceil(length_value))
                    if rounded_value > 0:
                        row.append(rounded_value)
                    else:
                        row.append("")
            # Append totals/perimeter columns
            rounded_total = int(math.ceil(fascade_assigned_totals.get(floor_id, 0.0)))
            row.append(rounded_total)
            perimeter_value = int(math.ceil(fascade_perimeters.get(floor_id, 0.0)))
            row.append(perimeter_value)
            fascade_sheet.append(row)
        
        if fascade_floor_numbers:
            for row in fascade_sheet.iter_rows(
                min_row=2,
                max_row=len(fascade_floor_numbers) + 1,
                min_col=1,
                max_col=1
            ):
                for cell in row:
                    cell.font = Font(bold=True)
        
        for row in fascade_sheet.iter_rows(
            min_row=2,
            max_row=len(fascade_floor_numbers) + 1,
            min_col=2
        ):
            for cell in row:
                if cell.value not in ("", None):
                    cell.alignment = right_align
        
        fascade_sheet.column_dimensions['A'].width = 15
        for col_idx, column_label in enumerate(fascade_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            label_width = len(str(column_label))
            fascade_sheet.column_dimensions[col_letter].width = max(label_width + 2, 8)
        # Set widths for total/perimeter columns
        total_col = openpyxl.utils.get_column_letter(len(fascade_labels) + 2)
        perimeter_col = openpyxl.utils.get_column_letter(len(fascade_labels) + 3)
        fascade_sheet.column_dimensions[total_col].width = 18
        fascade_sheet.column_dimensions[perimeter_col].width = 16
        
        fascade_sheet.freeze_panes = 'B2'
        
        # Print coverage summary
        print("\nFascade coverage summary:")
        for floor_id in fascade_floor_numbers:
            perimeter_ft = fascade_perimeters.get(floor_id, 0.0)
            assigned_ft = fascade_assigned_totals.get(floor_id, 0.0)
            threshold_used = fascade_thresholds.get(floor_id, FASCADE_DISTANCE_THRESHOLD)
            coverage_ratio = fascade_coverage.get(floor_id, 0.0)
            max_distance_seen = fascade_max_distances.get(floor_id, 0.0)
            method = fascade_methods.get(floor_id, 'nearest_boundary_sample')
            coverage_pct = coverage_ratio * 100.0
            if perimeter_ft > 1e-6:
                print(
                    f"  {floor_id}: {assigned_ft:.1f} ft assigned / {perimeter_ft:.1f} ft perimeter "
                    f"({coverage_pct:.1f}% coverage, method {method}, threshold {threshold_used:.1f} ft, "
                    f"max gap {max_distance_seen:.1f} ft)"
                )
            else:
                print(f"  {floor_id}: no perimeter detected (threshold {threshold_used:.1f} ft)")
        
        # --- Master KLL Sheet ---
        kll_data = collect_kll_matrix_data(floor_plans)
        kll_labels = kll_data['column_labels']
        kll_floor_numbers = kll_data['floor_numbers']
        kll_matrix = kll_data['matrix']

        kll_sheet = workbook.create_sheet(title="MASTER KLL")

        kll_header = ["Slab #"] + kll_labels
        kll_sheet.append(kll_header)
        for cell in kll_sheet[1]:
            cell.font = header_font

        for floor_id in kll_floor_numbers:
            row_data = [floor_id]
            for label in kll_labels:
                kll_value = kll_matrix[floor_id].get(label)
                row_data.append(kll_value if kll_value is not None else "")
            kll_sheet.append(row_data)

        if kll_floor_numbers:
            for row in kll_sheet.iter_rows(
                min_row=2, max_row=len(kll_floor_numbers) + 1,
                min_col=1, max_col=1
            ):
                for cell in row:
                    cell.font = Font(bold=True)

        for row in kll_sheet.iter_rows(
            min_row=2, max_row=len(kll_floor_numbers) + 1, min_col=2
        ):
            for cell in row:
                if cell.value not in ("", None):
                    cell.alignment = right_align

        kll_sheet.column_dimensions['A'].width = 15
        for col_idx, label in enumerate(kll_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            label_width = len(str(label))
            kll_sheet.column_dimensions[col_letter].width = max(label_width + 2, 8)

        kll_sheet.freeze_panes = 'B2'

        # Save Excel workbook
        workbook.save(output_filename)

        # Log success message with filename
        print(f"\n✓ Excel export successful: {output_filename}")
        print(f"  Master matrix: {len(floor_numbers)} floors × {len(column_labels)} columns")
        if add_labels:
            print(f"  Master additional load area sheet: {len(add_floor_numbers)} floors × {len(add_labels)} columns")
        if xs_labels:
            print(f"  Master cross section sheet: {len(xs_floor_numbers)} floors × {len(xs_labels)} columns")
        if fascade_labels:
            print(f"  Fascade length sheet: {len(fascade_floor_numbers)} floors × {len(fascade_labels)} boundary participants")
        else:
            print("  Fascade length sheet: no qualifying boundary participants (sheet contains floors only)")
        if kll_labels:
            print(f"  Master KLL sheet: {len(kll_floor_numbers)} floors × {len(kll_labels)} columns")
        if load_zone_rows:
            load_zone_names = sorted({str(row['load_zone']) for row in load_zone_rows})
            print(
                f"  Load zone area sheet: {len(load_zone_rows)} rows across "
                f"{len(load_zone_names)} load zone(s)"
            )
        
    except PermissionError:
        # Handle file permission errors
        print(f"\n✗ ERROR: Could not save Excel file '{output_filename}'")
        print(f"  The file may be open in another program.")
        print(f"  Please close the file and try again, or use a different filename.")
        print(f"  Suggested alternative: column_load_takedown_new.xlsx")
        
    except Exception as e:
        # Handle other errors
        import traceback
        print(f"\n✗ ERROR: Excel export failed: {e}")
        print(f"  Suggested alternative filename: column_load_takedown_backup.xlsx")
        # Print traceback for debugging
        traceback.print_exc()
