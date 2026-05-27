"""
Excel export module for column load takedown data.

This module exports column tributary area data to Excel format for structural
engineering load takedown calculations in a master matrix format.
"""

import math
import re

import openpyxl
from openpyxl.styles import Alignment, Font

from fascade_utils import (
    compute_fascade_assignments,
    FASCADE_DISTANCE_THRESHOLD,
    FASCADE_SAMPLE_SPACING,
)


RANGE_FLOOR_RE = re.compile(r'^\s*([A-Za-z]*)(\d+)\s*[-–—]\s*([A-Za-z]*)(\d+)\s*$')


def alphanumeric_sort_key(label):
    """
    Sort key function for alphanumeric sorting.
    Handles labels like: 1, 2, 10, 23, A1, B2, UNLABELED_0, etc.
    """
    try:
        # Convert label to string if it isn't already
        if isinstance(label, bytes):
            label = label.decode('utf-8', errors='replace')
        elif not isinstance(label, str):
            label = str(label) if label is not None else ""
        
        # Split label into numeric and non-numeric parts
        parts = re.split(r'(\d+)', label)
        # Convert numeric parts to integers for proper sorting
        return [int(part) if part.isdigit() else part.lower() for part in parts]
    except Exception:
        # Fallback to simple string sorting if there's an error
        return [str(label)]


def floor_sort_key(floor_id):
    """
    Sort key function for floor identifiers in descending order.
    Handles: MAIN ROOF, ROOF, PENTHOUSE, numbered floors, GROUND, BASEMENT, etc.
    """
    floor_str = str(floor_id).upper()
    
    # Define priority order (higher number = appears first)
    if 'ROOF' in floor_str and 'MAIN' in floor_str:
        return (1000, floor_str)
    elif 'ROOF' in floor_str:
        return (900, floor_str)
    elif 'PENTHOUSE' in floor_str or 'PH' in floor_str:
        return (800, floor_str)
    elif 'GROUND' in floor_str or 'MAIN' in floor_str or 'LOBBY' in floor_str:
        return (-100, floor_str)
    elif 'BASEMENT' in floor_str or floor_str.startswith('B'):
        # Extract basement number if present
        match = re.search(r'B?(\d+)', floor_str)
        if match:
            basement_num = int(match.group(1))
            return (-200 - basement_num, floor_str)
        return (-200, floor_str)
    else:
        # Try to extract numeric floor number
        match = re.search(r'(\d+)', floor_str)
        if match:
            floor_num = int(match.group(1))
            return (floor_num, floor_str)
        # Default: alphabetical
        return (0, floor_str)


def expand_floor_identifier(floor_id):
    """
    Expand grouped floor labels such as ``4-8`` or ``B1-B3`` into
    individual floor identifiers for spreadsheet output.
    """
    floor_str = str(floor_id).strip()
    match = RANGE_FLOOR_RE.fullmatch(floor_str)
    if not match:
        return [floor_str]

    start_prefix, start_num, end_prefix, end_num = match.groups()
    if start_prefix.upper() != end_prefix.upper():
        return [floor_str]

    start_value = int(start_num)
    end_value = int(end_num)
    step = 1 if end_value >= start_value else -1
    prefix = start_prefix.upper()
    return [f"{prefix}{value}" if prefix else str(value) for value in range(start_value, end_value + step, step)]


def expand_floor_map(source_floor_data):
    """
    Duplicate grouped floor data into one spreadsheet row per represented floor.
    """
    expanded = {}
    for floor_id, value_map in source_floor_data.items():
        expanded_floor_ids = expand_floor_identifier(floor_id)
        for expanded_floor_id in expanded_floor_ids:
            if expanded_floor_id not in expanded:
                expanded[expanded_floor_id] = {}
            expanded[expanded_floor_id].update(value_map)
    return expanded


def expand_floor_scalar_map(source_values):
    """Duplicate per-floor scalar values for grouped floor identifiers."""
    expanded = {}
    for floor_id, value in source_values.items():
        for expanded_floor_id in expand_floor_identifier(floor_id):
            expanded[expanded_floor_id] = value
    return expanded


def collect_master_matrix_data(floor_plans):
    """
    Collect and organize data for master matrix format.
    
    Args:
        floor_plans: List of floor plan dictionaries with column data
        
    Returns:
        dict with keys:
            'floor_numbers': List of floor identifiers (sorted descending)
            'column_labels': List of unique column labels (sorted ascending)
            'matrix': Dict[floor][column] = tributary_area (rounded up)
    """
    # Collect all unique column labels across all floor plans
    all_column_labels = set()
    
    # Build a mapping: floor_id -> {column_label -> tributary_area}
    floor_data = {}
    
    for floor_plan in floor_plans:
        # Get floor identifier (use floor_number if available, otherwise boundary_id)
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        
        column_points = floor_plan.get('column_points', [])
        column_labels = floor_plan.get('column_labels', [])
        areas = floor_plan.get('areas', [])
        point_metadata = floor_plan.get('point_metadata', [])
        
        # Initialize floor data
        if floor_id not in floor_data:
            floor_data[floor_id] = {}
        
        if point_metadata:
            # Legacy mixed-site shape: iterate through metadata to extract column tributary areas.
            for i, metadata in enumerate(point_metadata):
                if metadata.get('type') != 'column':
                    continue

                col_idx = metadata.get('column_index')
                if col_idx is None or col_idx >= len(column_points):
                    continue

                if col_idx < len(column_labels):
                    label = column_labels[col_idx]
                    if not isinstance(label, str):
                        label = str(label) if label is not None else f"UNLABELED_{col_idx}"
                else:
                    label = f"UNLABELED_{col_idx}"

                tributary_area = areas[i] if i < len(areas) else 0.0
                rounded_area = math.ceil(tributary_area)
                floor_data[floor_id][label] = rounded_area
                all_column_labels.add(label)
        else:
            # Column-only shape: `areas` already aligns with `column_points` / `column_labels`.
            for col_idx, point in enumerate(column_points):
                label = column_labels[col_idx] if col_idx < len(column_labels) else f"UNLABELED_{col_idx}"
                if not isinstance(label, str):
                    label = str(label) if label is not None else f"UNLABELED_{col_idx}"

                tributary_area = areas[col_idx] if col_idx < len(areas) else 0.0
                rounded_area = math.ceil(tributary_area)
                floor_data[floor_id][label] = rounded_area
                all_column_labels.add(label)
    
    floor_data = expand_floor_map(floor_data)

    # Sort column labels alphanumerically (ascending: 1, 2, 3, ..., N)
    sorted_column_labels = sorted(list(all_column_labels), key=alphanumeric_sort_key)
    
    # Sort floor numbers in descending order (MAIN ROOF → 2ND → 1ST)
    sorted_floor_numbers = sorted(list(floor_data.keys()), key=floor_sort_key, reverse=True)
    
    # Build matrix data structure: matrix[floor][column] = tributary_area
    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for column_label in sorted_column_labels:
            # Get tributary area if column exists on this floor, otherwise None
            matrix[floor_id][column_label] = floor_data[floor_id].get(column_label, None)
    
    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_column_labels,
        'matrix': matrix
    }


def collect_fascade_length_data(
    floor_plans,
    distance_threshold=FASCADE_DISTANCE_THRESHOLD,
    sample_spacing=FASCADE_SAMPLE_SPACING
):
    """Aggregate façade length attribution data for every floor."""
    floor_data = {}
    all_labels = set()
    floor_perimeters = {}
    floor_assigned_totals = {}
    floor_thresholds = {}
    floor_coverages = {}
    floor_max_distances = {}
    floor_methods = {}
    
    for floor_plan in floor_plans:
        floor_id = floor_plan.get('floor_number', floor_plan.get('boundary_id', 'UNKNOWN'))
        if floor_plan.get('fascade_data'):
            result = floor_plan['fascade_data']
        else:
            result = compute_fascade_assignments(
                floor_plan,
                distance_threshold=distance_threshold,
                sample_spacing=sample_spacing
            )
            floor_plan['fascade_data'] = result
        
        length_map = result.get('length_map', {})
        floor_perimeters[floor_id] = result.get('perimeter', 0.0)
        floor_assigned_totals[floor_id] = result.get('assigned_total', 0.0)
        floor_thresholds[floor_id] = result.get('threshold_used', distance_threshold)
        floor_coverages[floor_id] = result.get('coverage_ratio', 0.0)
        floor_max_distances[floor_id] = result.get('max_distance_seen', 0.0)
        floor_methods[floor_id] = result.get('assignment_method', 'nearest_boundary_sample')
        
        if floor_id not in floor_data:
            floor_data[floor_id] = {}
        
        for label, length in length_map.items():
            floor_data[floor_id][label] = floor_data[floor_id].get(label, 0.0) + length
            all_labels.add(label)
    
    floor_data = expand_floor_map(floor_data)
    floor_perimeters = expand_floor_scalar_map(floor_perimeters)
    floor_assigned_totals = expand_floor_scalar_map(floor_assigned_totals)
    floor_thresholds = expand_floor_scalar_map(floor_thresholds)
    floor_coverages = expand_floor_scalar_map(floor_coverages)
    floor_max_distances = expand_floor_scalar_map(floor_max_distances)
    floor_methods = expand_floor_scalar_map(floor_methods)

    sorted_floor_numbers = sorted(list(floor_data.keys()), key=floor_sort_key, reverse=True)
    sorted_labels = sorted(list(all_labels), key=alphanumeric_sort_key)
    
    matrix = {}
    for floor_id in sorted_floor_numbers:
        matrix[floor_id] = {}
        for label in sorted_labels:
            matrix[floor_id][label] = floor_data.get(floor_id, {}).get(label)
    
    return {
        'floor_numbers': sorted_floor_numbers,
        'column_labels': sorted_labels,
        'matrix': matrix,
        'perimeters': floor_perimeters,
        'assigned_totals': floor_assigned_totals,
        'thresholds': floor_thresholds,
        'coverage': floor_coverages,
        'max_distances': floor_max_distances,
        'methods': floor_methods,
    }


def export_column_load_takedown(floor_plans, output_filename="column_load_takedown.xlsx"):
    """
    Export column load takedown data to Excel in master matrix format.
    
    Args:
        floor_plans: List of floor plan dictionaries with column data
        output_filename: Path to output Excel file
        
    Returns:
        None (writes file to disk)
    """
    try:
        # Collect master matrix data
        matrix_data = collect_master_matrix_data(floor_plans)
        
        floor_numbers = matrix_data['floor_numbers']
        column_labels = matrix_data['column_labels']
        matrix = matrix_data['matrix']
        
        # Create new Workbook object
        workbook = openpyxl.Workbook()
        
        # Remove the default sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create single worksheet named "MASTER TRIBUTARY AREA"
        worksheet = workbook.create_sheet(title="MASTER TRIBUTARY AREA")
        
        # Write header row: "Floor" in first column, then column labels
        header_row = ["Floor"] + column_labels
        worksheet.append(header_row)
        
        # Apply bold formatting to header row
        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font
        
        # Write floor rows
        for floor_id in floor_numbers:
            # Start with floor identifier in first column
            row_data = [floor_id]
            
            # Add tributary areas for each column
            for column_label in column_labels:
                area = matrix[floor_id].get(column_label)
                # Leave cell empty if column doesn't exist on this floor
                row_data.append(area if area is not None else "")
            
            worksheet.append(row_data)
        
        # Apply formatting
        # Bold first column (floor numbers)
        for row in worksheet.iter_rows(min_row=2, max_row=len(floor_numbers) + 1, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)
        
        # Right-align numeric cells (all columns except first)
        right_align = Alignment(horizontal='right')
        for row in worksheet.iter_rows(min_row=2, max_row=len(floor_numbers) + 1, min_col=2):
            for cell in row:
                if cell.value and cell.value != "":
                    cell.alignment = right_align
        
        # Set column widths
        # First column (Floor) - wider for floor names
        worksheet.column_dimensions['A'].width = 15
        
        # Other columns - auto-size based on content
        for col_idx, column_label in enumerate(column_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            # Calculate width based on label length and typical area values
            label_width = len(str(column_label))
            worksheet.column_dimensions[col_letter].width = max(label_width + 2, 8)
        
        # Freeze top row and first column for scrolling
        worksheet.freeze_panes = 'B2'
        
        # --- Fascade Length Sheet ---
        fascade_data = collect_fascade_length_data(floor_plans)
        fascade_labels = fascade_data['column_labels']
        fascade_floor_numbers = fascade_data['floor_numbers']
        fascade_matrix = fascade_data['matrix']
        fascade_perimeters = fascade_data.get('perimeters', {})
        fascade_assigned_totals = fascade_data.get('assigned_totals', {})
        fascade_thresholds = fascade_data.get('thresholds', {})
        fascade_coverage = fascade_data.get('coverage', {})
        fascade_max_distances = fascade_data.get('max_distances', {})
        fascade_methods = fascade_data.get('methods', {})
        
        fascade_sheet = workbook.create_sheet(title="FASCADE LENGTH")
        fascade_header = ["Floor"] + fascade_labels + ["TOTAL ASSIGNED (FT)", "PERIMETER (FT)"]
        fascade_sheet.append(fascade_header)
        for cell in fascade_sheet[1]:
            cell.font = header_font
        
        for floor_id in fascade_floor_numbers:
            row = [floor_id]
            floor_lengths = fascade_matrix.get(floor_id, {})
            for label in fascade_labels:
                length_value = floor_lengths.get(label)
                if length_value is None:
                    row.append("")
                else:
                    rounded_value = int(math.ceil(length_value))
                    if rounded_value > 0:
                        row.append(rounded_value)
                    else:
                        row.append("")
            # Append totals/perimeter columns
            rounded_total = int(math.ceil(fascade_assigned_totals.get(floor_id, 0.0)))
            row.append(rounded_total)
            perimeter_value = int(math.ceil(fascade_perimeters.get(floor_id, 0.0)))
            row.append(perimeter_value)
            fascade_sheet.append(row)
        
        if fascade_floor_numbers:
            for row in fascade_sheet.iter_rows(
                min_row=2,
                max_row=len(fascade_floor_numbers) + 1,
                min_col=1,
                max_col=1
            ):
                for cell in row:
                    cell.font = Font(bold=True)
        
        for row in fascade_sheet.iter_rows(
            min_row=2,
            max_row=len(fascade_floor_numbers) + 1,
            min_col=2
        ):
            for cell in row:
                if cell.value not in ("", None):
                    cell.alignment = right_align
        
        fascade_sheet.column_dimensions['A'].width = 15
        for col_idx, column_label in enumerate(fascade_labels, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            label_width = len(str(column_label))
            fascade_sheet.column_dimensions[col_letter].width = max(label_width + 2, 8)
        # Set widths for total/perimeter columns
        total_col = openpyxl.utils.get_column_letter(len(fascade_labels) + 2)
        perimeter_col = openpyxl.utils.get_column_letter(len(fascade_labels) + 3)
        fascade_sheet.column_dimensions[total_col].width = 18
        fascade_sheet.column_dimensions[perimeter_col].width = 16
        
        fascade_sheet.freeze_panes = 'B2'
        
        # Print coverage summary
        print("\nFascade coverage summary:")
        for floor_id in fascade_floor_numbers:
            perimeter_ft = fascade_perimeters.get(floor_id, 0.0)
            assigned_ft = fascade_assigned_totals.get(floor_id, 0.0)
            threshold_used = fascade_thresholds.get(floor_id, FASCADE_DISTANCE_THRESHOLD)
            coverage_ratio = fascade_coverage.get(floor_id, 0.0)
            max_distance_seen = fascade_max_distances.get(floor_id, 0.0)
            method = fascade_methods.get(floor_id, 'nearest_boundary_sample')
            coverage_pct = coverage_ratio * 100.0
            if perimeter_ft > 1e-6:
                print(
                    f"  {floor_id}: {assigned_ft:.1f} ft assigned / {perimeter_ft:.1f} ft perimeter "
                    f"({coverage_pct:.1f}% coverage, method {method}, threshold {threshold_used:.1f} ft, "
                    f"max gap {max_distance_seen:.1f} ft)"
                )
            else:
                print(f"  {floor_id}: no perimeter detected (threshold {threshold_used:.1f} ft)")
        
        # Save Excel workbook
        workbook.save(output_filename)
        
        # Log success message with filename
        print(f"\n✓ Excel export successful: {output_filename}")
        print(f"  Master matrix: {len(floor_numbers)} floors × {len(column_labels)} columns")
        if fascade_labels:
            print(f"  Fascade length sheet: {len(fascade_floor_numbers)} floors × {len(fascade_labels)} boundary participants")
        else:
            print("  Fascade length sheet: no qualifying boundary participants (sheet contains floors only)")
        
    except PermissionError:
        # Handle file permission errors
        print(f"\n✗ ERROR: Could not save Excel file '{output_filename}'")
        print(f"  The file may be open in another program.")
        print(f"  Please close the file and try again, or use a different filename.")
        print(f"  Suggested alternative: column_load_takedown_new.xlsx")
        
    except Exception as e:
        # Handle other errors
        import traceback
        print(f"\n✗ ERROR: Excel export failed: {e}")
        print(f"  Suggested alternative filename: column_load_takedown_backup.xlsx")
        # Print traceback for debugging
        traceback.print_exc()
